import openpyxl
import re
import os
import urllib.request
import json
import time

EXCEL_PATH = 'bin/Legal_Matching_Pok balls.xlsx'
CUSTOM_JSON_PATH = 'bin/custom_matching_balls.json'
OUTPUT_PATH = 'lib/data/matching_balls_data.dart'

TYPE_BALL_MAPPING = {
    'normal': 'premier_ball',
    'fire': 'fast_ball',
    'water': 'lure_ball',
    'electric': 'quick_ball',
    'grass': 'friend_ball',
    'ice': 'dive_ball',
    'fighting': 'level_ball',
    'poison': 'timer_ball',
    'ground': 'safari_ball',
    'flying': 'repeat_ball',
    'psychic': 'dream_ball',
    'bug': 'net_ball',
    'rock': 'ultra_ball',
    'ghost': 'dusk_ball',
    'dragon': 'great_ball',
    'dark': 'luxury_ball',
    'steel': 'heavy_ball',
    'fairy': 'love_ball'
}

if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = 'Legal_Matching_Pok balls.xlsx'

def get_url(cell_value):
    if not isinstance(cell_value, str):
        return None
    match = re.search(r'(https?://[^\s"<>&\u0027\)]+)', cell_value)
    return match.group(1) if match else None

def get_form(name, is_direct=False):
    name = str(name).lower().replace('é', 'e').replace('è', 'e').strip()
    
    if is_direct:
        return re.sub(r'\s+', '-', name).replace("'", "")
        
    if 'alola' in name: return 'alola'
    if 'galar' in name: return 'galar'
    if 'hisui' in name: return 'hisui'
    if 'paldea' in name: return 'paldea'
    
    if 'mega' in name:
        if ' x' in name: return 'mega-x'
        if ' y' in name: return 'mega-y'
        return 'mega'
        
    if 'primal' in name: return 'primal'
    if 'gmax' in name or 'gigantamax' in name: return 'gmax'
    
    if '(' in name:
        form_part = name.split('(')[1].replace(')', '').strip()
        form_part = form_part.lower().replace('é', 'e')
        
        base_forms = [
            'no plate', 'altered', 'land', 'incarnate', 'aria', 'ordinary',
            'shield', '50%', 'confined', 'baile', 'midday', 'solo', 'red core',
            'disguised', 'amped', 'ice face', 'full belly', 'single strike',
            'hero', 'chest', 'family of four', 'green plumage', 'curly', 'plant', 'west',
            'spring', 'normal', 'base', 'standard', 'red-striped', 'red meteor', 'red',
            'natural'
        ]
        if form_part in base_forms:
            return 'normal'
            
        return re.sub(r'\s+', '-', form_part).replace("'", "")
        
    return 'normal'

def to_key(name):
    return name.lower().replace(' ', 'e').replace(' ', '_').strip()

def main():
    actual_path = EXCEL_PATH
    if not os.path.exists(actual_path):
        for file in os.listdir('.'):
            if 'Legal_Matching' in file and file.endswith('.xlsx'):
                actual_path = file
                break
        for file in os.listdir('bin'):
            if 'Legal_Matching' in file and file.endswith('.xlsx'):
                actual_path = f"bin/{file}"
                break

    print(f"Starte Parsing der Excel-Datei: {actual_path} ...")
    wb = openpyxl.load_workbook(actual_path, data_only=False)
    
    url_to_key = {}
    key_to_url = {}
    database = {}
    
    intro_sheet = wb['Intro']
    for row in intro_sheet.iter_rows(min_row=1, max_row=100, values_only=True):
        for c in range(len(row) - 1):
            url = get_url(row[c])
            if url:
                name = str(row[c+1]).replace('"', '').strip()
                if 'ball' in name.lower():
                    key = to_key(name)
                    url_to_key[url] = key
                    key_to_url[key] = url

    url_to_key['https://i.imgur.com/eru43o1.png'] = 'strange_ball'
    key_to_url['strange_ball'] = 'https://i.imgur.com/eru43o1.png'
    url_to_key['https://i.imgur.com/aeqHLEh.png'] = 'cherish_ball'
    key_to_url['cherish_ball'] = 'https://i.imgur.com/aeqHLEh.png'

    print(f"  Mapping erstellt: {len(url_to_key)} Bälle gefunden.")
    
    for sheet_name in wb.sheetnames:
        if sheet_name in ['Intro', 'Vivillon', 'Alcremie']:
            continue
            
        print(f"Lese Tab: {sheet_name} ...")
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        entries = 0
        
        for r in range(0, len(rows), 7):
            if r >= len(rows): break
            row_0 = rows[r]
            
            for c in range(len(row_0) - 1):
                id_val = str(row_0[c] or '').strip()
                id_match = re.sub(r'[^0-9]', '', id_val)
                if not id_match: continue
                
                poke_id = int(id_match)
                name_str = str(row_0[c+1] or '').replace('*', '').strip()
                form = get_form(name_str)
                unique_id = f"{poke_id}_{form}".replace("'", "")
                
                normal_balls = []
                for offset in [1, 2, 3]:
                    if r + offset < len(rows) and c < len(rows[r+offset]):
                        url = get_url(rows[r+offset][c])
                        if url and url in url_to_key and url_to_key[url] not in normal_balls:
                            normal_balls.append(url_to_key[url])
                
                shiny_balls = []
                for offset in [4, 5, 6]:
                    if r + offset < len(rows) and c < len(rows[r+offset]):
                        url = get_url(rows[r+offset][c])
                        if url and url in url_to_key and url_to_key[url] not in shiny_balls:
                            shiny_balls.append(url_to_key[url])
                
                norm_str = "['" + "', '".join(normal_balls) + "']" if normal_balls else "['any_ball']"
                shin_str = "['" + "', '".join(shiny_balls) + "']" if shiny_balls else "['any_ball']"
                
                database[unique_id] = {'normal': norm_str, 'shiny': shin_str}
                entries += 1
        print(f"  -> {entries} Einträge gefunden.")
        
    print("Lese Tab: Vivillon ...")
    if 'Vivillon' in wb.sheetnames:
        viv_sheet = wb['Vivillon']
        viv_rows = list(viv_sheet.iter_rows(values_only=True))
        entries = 0
        vivillon_forms = [
            'meadow', 'icy-snow', 'polar', 'tundra', 'continental', 'garden',
            'elegant', 'modern', 'marine', 'archipelago', 'high-plains',
            'sandstorm', 'river', 'monsoon', 'savanna', 'sun', 'ocean',
            'jungle', 'fancy', 'poke-ball'
        ]
        
        for r in range(len(viv_rows)):
            row_data = viv_rows[r]
            for c in range(len(row_data)):
                cell_val = str(row_data[c] or '').lower().strip()
                if not cell_val or len(cell_val) < 3 or 'example' in cell_val or 'note' in cell_val: continue
                
                detected_form = None
                for v_form in vivillon_forms:
                    if v_form.replace('-', ' ') in cell_val.replace('-', ' ') or v_form.replace('-', '') in cell_val.replace('-', ''):
                        detected_form = v_form
                        break
                
                if detected_form:
                    unique_id = f"666_{detected_form}"
                    normal_balls, shiny_balls = [], []
                    
                    for r_offset in range(1, 4):
                        if r + r_offset < len(viv_rows):
                            search_row = viv_rows[r + r_offset]
                            for c_offset in [c-1, c, c+1]:
                                if 0 <= c_offset < len(search_row):
                                    url = get_url(search_row[c_offset])
                                    if url and url in url_to_key:
                                        if r_offset == 1 and url_to_key[url] not in normal_balls:
                                            normal_balls.append(url_to_key[url])
                                        elif r_offset >= 2 and url_to_key[url] not in shiny_balls:
                                            shiny_balls.append(url_to_key[url])
                    
                    norm_str = "['" + "', '".join(normal_balls) + "']" if normal_balls else "['any_ball']"
                    shin_str = "['" + "', '".join(shiny_balls) + "']" if shiny_balls else (norm_str if normal_balls else "['any_ball']")
                    
                    if normal_balls or unique_id not in database:
                        database[unique_id] = {'normal': norm_str, 'shiny': shin_str}
                        entries += 1
        print(f"  -> {entries} Einträge gefunden.")

    print("Lese Tab: Alcremie ...")
    if 'Alcremie' in wb.sheetnames:
        alc_sheet = wb['Alcremie']
        alc_rows = list(alc_sheet.iter_rows(values_only=True))
        entries = 0
        for r in range(len(alc_rows)):
            row_data = alc_rows[r]
            if not row_data or not row_data[0]: continue
            
            cell_val = str(row_data[0]).lower().strip()
            if not cell_val or len(cell_val) < 4 or 'note' in cell_val: continue
            
            form = None
            if 'ruby' in cell_val and 'swirl' in cell_val: form = 'ruby-swirl'
            elif 'ruby' in cell_val: form = 'ruby-cream'
            elif 'caramel' in cell_val: form = 'caramel-swirl'
            elif 'rainbow' in cell_val: form = 'rainbow-swirl'
            elif 'vanilla' in cell_val: form = 'vanilla-cream'
            elif 'matcha' in cell_val: form = 'matcha-cream'
            elif 'mint' in cell_val: form = 'mint-cream'
            elif 'lemon' in cell_val: form = 'lemon-cream'
            elif 'salted' in cell_val: form = 'salted-cream'
            
            if form:
                unique_id = f"869_{form}"
                normal_balls, shiny_balls = [], []
                
                for c in range(1, len(row_data)):
                    url = get_url(row_data[c])
                    if url and url in url_to_key and url_to_key[url] not in normal_balls:
                        normal_balls.append(url_to_key[url])
                
                if r + 1 < len(alc_rows):
                    row_shiny = alc_rows[r+1]
                    for c in range(1, len(row_shiny)):
                        url = get_url(row_shiny[c])
                        if url and url in url_to_key and url_to_key[url] not in shiny_balls:
                            shiny_balls.append(url_to_key[url])
                
                norm_str = "['" + "', '".join(normal_balls) + "']" if normal_balls else "['any_ball']"
                shin_str = "['" + "', '".join(shiny_balls) + "']" if shiny_balls else "['any_ball']"
                
                if normal_balls or unique_id not in database:
                    database[unique_id] = {'normal': norm_str, 'shiny': shin_str}
                    entries += 1
        print(f"  -> {entries} Einträge gefunden.")

    print("\nLese benutzerdefinierte Bälle aus JSON (Community Fixes)...")
    if os.path.exists(CUSTOM_JSON_PATH):
        try:
            with open(CUSTOM_JSON_PATH, 'r', encoding='utf-8') as f:
                custom_data = json.load(f)
                custom_entries = 0
                for unique_id, data in custom_data.items():
                    norm_str = "['" + "', '".join(data.get('normal', [])) + "']" if data.get('normal') else "['any_ball']"
                    shin_str = "['" + "', '".join(data.get('shiny', [])) + "']" if data.get('shiny') else "['any_ball']"
                    
                    database[unique_id] = {'normal': norm_str, 'shiny': shin_str}
                    custom_entries += 1
            print(f"  -> {custom_entries} benutzerdefinierte Einträge aus {CUSTOM_JSON_PATH} geladen und überschrieben.")
        except Exception as e:
            print(f"  -> Fehler beim Laden der JSON: {e}")
    else:
        print(f"  -> Keine {CUSTOM_JSON_PATH} gefunden. Überspringe Community Fixes.")

    print("\nFühre API Fallbacks für fehlende Bälle (Zwischenentwicklungen) durch...")
    root_cache = {}
    fallback_count = 0
    for unique_id, data in database.items():
        if data['normal'] == "['any_ball']":
            parts = unique_id.split('_', 1)
            if not parts[0].isdigit(): continue
            poke_id = int(parts[0])
            form = parts[1]
            
            curr_id = poke_id
            while curr_id not in root_cache:
                try:
                    req = urllib.request.Request(f"https://pokeapi.co/api/v2/pokemon-species/{curr_id}/", headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req) as response:
                        res_data = json.loads(response.read().decode())
                        if res_data.get('evolves_from_species'):
                            url = res_data['evolves_from_species']['url']
                            parent_id = int(url.rstrip('/').split('/')[-1])
                            curr_id = parent_id
                            time.sleep(0.05) 
                        else:
                            root_cache[curr_id] = curr_id
                            break
                except Exception:
                    root_cache[curr_id] = curr_id
                    break
            
            root_id = root_cache.get(curr_id, curr_id)
            root_cache[poke_id] = root_id
            
            if root_id != poke_id:
                base_key = f"{root_id}_{form}"
                if base_key not in database:
                    base_key = f"{root_id}_normal"
                
                if base_key in database and database[base_key]['normal'] != "['any_ball']":
                    database[unique_id]['normal'] = database[base_key]['normal']
                    if data['shiny'] == "['any_ball']":
                        database[unique_id]['shiny'] = database[base_key]['shiny']
                    fallback_count += 1
    print(f"  -> {fallback_count} Entwicklungen erfolgreich mit Bällen der Basis-Form befüllt!")

    print("\nSetze spezifische Bälle für Arceus (493) und Amigento (773) anhand ihrer Typen ...")
    for poke_id in [493, 773]:
        for type_name, ball_key in TYPE_BALL_MAPPING.items():
            form_id = f"{poke_id}_{type_name}"
            database[form_id] = {
                'normal': f"['{ball_key}']",
                'shiny': f"['{ball_key}']"
            }
            if type_name == 'normal':
                database[f"{poke_id}_normal"] = {
                    'normal': f"['{ball_key}']",
                    'shiny': f"['{ball_key}']"
                }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write("// Automatisch generierte Datei - NICHT MANUELL ÄNDERN!\n")
        f.write("// Quelle: Python Parser für \"Legal_Matching_Pokeballs.xlsx\"\n\n")
        
        f.write("const Map<String, String> ballImageUrls = {\n")
        for k, u in key_to_url.items():
            f.write(f"  '{k}': '{u}',\n")
        f.write("};\n\n")
        
        f.write("const Map<String, Map<String, List<String>>> matchingBallsDatabase = {\n")
        for k, v in database.items():
            f.write(f"  '{k}': {{'normal': {v['normal']}, 'shiny': {v['shiny']}}},\n")
        f.write("};\n")
        
    print(f"\nErfolgreich! {len(database)} Einträge in {OUTPUT_PATH} gespeichert.")

if __name__ == '__main__':
    main()