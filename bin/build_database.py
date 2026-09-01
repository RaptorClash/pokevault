import urllib.request
import json
import sqlite3
import os
import re
import time
import openpyxl

EXCEL_PATH = 'bin/Legal_Matching_Pokeballs.xlsx'
CUSTOM_JSON_PATH = 'bin/custom_matching_balls.json'

FORM_WHITELIST = {
    "172_spiky-eared": ["updated_johto_regional"],
    "25_cosplay": ["updated_hoenn_regional"],
    "25_rockstar": ["updated_hoenn_regional"],
    "25_belle": ["updated_hoenn_regional"],
    "25_popstar": ["updated_hoenn_regional"],
    "25_phd": ["updated_hoenn_regional"],
    "25_libre": ["updated_hoenn_regional"],
    "25_starter": ["kanto_regional"],
    "133_starter": ["kanto_regional"],
    "670_eternal": ["kalos_central_regional", "kalos_coastal_regional", "kalos_mountain_regional", "lumiose_regional"],
}

VERSION_TO_GEN = {
    'red': 'gen_1', 'blue': 'gen_1', 'yellow': 'gen_1',
    'gold': 'gen_2', 'silver': 'gen_2', 'crystal': 'gen_2',
    'ruby': 'gen_3', 'sapphire': 'gen_3', 'emerald': 'gen_3', 'firered': 'gen_3', 'leafgreen': 'gen_3', 'colosseum': 'gen_3', 'xd': 'gen_3',
    'diamond': 'gen_4', 'pearl': 'gen_4', 'platinum': 'gen_4', 'heartgold': 'gen_4', 'soulsilver': 'gen_4',
    'black': 'gen_5', 'white': 'gen_5', 'black-2': 'gen_5', 'white-2': 'gen_5',
    'x': 'gen_6', 'y': 'gen_6', 'omega-ruby': 'gen_6', 'alpha-sapphire': 'gen_6',
    'sun': 'gen_7', 'moon': 'gen_7', 'ultra-sun': 'gen_7', 'ultra-moon': 'gen_7', 'lets-go-pikachu': 'gen_7', 'lets-go-eevee': 'gen_7',
    'sword': 'gen_8', 'shield': 'gen_8', 'the-isle-of-armor-sword': 'gen_8', 'the-isle-of-armor-shield': 'gen_8', 'the-crown-tundra-sword': 'gen_8', 'the-crown-tundra-shield': 'gen_8', 'brilliant-diamond': 'gen_8', 'shining-pearl': 'gen_8', 'legends-arceus': 'gen_8',
    'scarlet': 'gen_9', 'violet': 'gen_9', 'the-teal-mask-scarlet': 'gen_9', 'the-teal-mask-violet': 'gen_9', 'the-indigo-disk-scarlet': 'gen_9', 'the-indigo-disk-violet': 'gen_9', 'legends-z-a': 'gen_9',
}

HARDCODED_ORDERS = {
    'lumiose': [152, 153, 154, 498, 499, 500, 158, 159, 160, 661, 662, 663, 659, 660, 664, 665, 666, 13, 14, 15, 16, 17, 18, 179, 180, 181, 504, 505, 406, 315, 407, 129, 130, 688, 689, 120, 121, 669, 670, 671, 672, 673, 677, 678, 667, 668, 674, 675, 568, 569, 702, 172, 25, 26, 173, 35, 36, 167, 168, 23, 24, 63, 64, 65, 92, 93, 94, 543, 544, 545, 679, 680, 681, 69, 70, 71, 511, 512, 513, 514, 515, 516, 307, 308, 309, 310, 280, 281, 282, 475, 228, 229, 333, 334, 531, 682, 683, 684, 685, 133, 134, 135, 136, 196, 197, 470, 471, 700, 427, 428, 353, 354, 582, 583, 584, 322, 323, 449, 450, 529, 530, 551, 552, 553, 66, 67, 68, 443, 444, 445, 703, 302, 303, 359, 447, 448, 79, 80, 199, 318, 319, 602, 603, 604, 147, 148, 149, 1, 2, 3, 4, 5, 6, 7, 8, 9, 618, 676, 686, 687, 690, 691, 692, 693, 704, 705, 706, 225, 361, 362, 478, 459, 460, 712, 713, 123, 212, 127, 214, 587, 701, 708, 709, 559, 560, 714, 715, 707, 607, 608, 609, 142, 696, 697, 698, 699, 95, 208, 304, 305, 306, 694, 695, 710, 711, 246, 247, 248, 656, 657, 658, 870, 650, 651, 652, 227, 653, 654, 655, 371, 372, 373, 115, 780, 374, 375, 376, 716, 717, 718, 719, 150],
    'mega-dex': [3, 6, 9, 15, 18, 65, 80, 94, 115, 127, 130, 142, 150, 181, 208, 212, 214, 229, 248, 254, 257, 260, 282, 302, 303, 306, 308, 310, 319, 323, 334, 354, 359, 362, 373, 376, 380, 381, 382, 383, 384, 428, 445, 448, 460, 475, 719],
    'icognito-dex': [201],
}

REGIONAL_ENDPOINTS = {
    'kanto': 'kanto_regional', 'letsgo-kanto': 'letsgo_kanto_regional', 'original-johto': 'johto_regional',
    'updated-johto': 'updated_johto_regional', 'hoenn': 'hoenn_regional', 'updated-hoenn': 'updated_hoenn_regional',
    'original-sinnoh': 'sinnoh_regional', 'extended-sinnoh': 'extended_sinnoh_regional', 'original-unova': 'unova_regional',
    'updated-unova': 'updated_unova_regional', 'kalos-central': 'kalos_central_regional', 'kalos-coastal': 'kalos_coastal_regional',
    'kalos-mountain': 'kalos_mountain_regional', 'original-alola': 'alola_regional', 'original-melemele': 'melemele_regional',
    'original-akala': 'akala_regional', 'original-ulaula': 'ulaula_regional', 'original-poni': 'poni_regional',
    'updated-alola': 'updated_alola_regional', 'updated-melemele': 'updated_melemele_regional', 'updated-akala': 'updated_akala_regional',
    'updated-ulaula': 'updated_ulaula_regional', 'updated-poni': 'updated_poni_regional', 'galar': 'galar_regional',
    'isle-of-armor': 'isle_of_armor_regional', 'crown-tundra': 'crown_tundra_regional', 'hisui': 'hisui_regional',
    'paldea': 'paldea_regional', 'kitakami': 'kitakami_regional', 'blueberry': 'blueberry_regional', 'lumiose': 'lumiose_regional',
    'lumiose-dimensions': 'lumiose_dimensions_regional', 'mega-dex': 'mega_dex', 'icognito-dex': 'icognito_dex',
}

TYPE_BALL_MAPPING = {
    'normal': 'premier_ball', 'fire': 'fast_ball', 'water': 'lure_ball', 'electric': 'quick_ball',
    'grass': 'friend_ball', 'ice': 'dive_ball', 'fighting': 'level_ball', 'poison': 'timer_ball',
    'ground': 'safari_ball', 'flying': 'repeat_ball', 'psychic': 'dream_ball', 'bug': 'net_ball',
    'rock': 'ultra_ball', 'ghost': 'dusk_ball', 'dragon': 'great_ball', 'dark': 'luxury_ball',
    'steel': 'heavy_ball', 'fairy': 'love_ball'
}

def fetch_json(url):
    for _ in range(3):
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
            with urllib.request.urlopen(req) as response:
                return json.loads(response.read().decode())
        except Exception:
            time.sleep(1)
    return None

def clean_location(raw_loc):
    loc = raw_loc.lower()
    loc = re.sub(r'-?area$', '', loc)
    loc = re.sub(r'-?(south|north|east|west)-towards.*$', '', loc)
    loc = re.sub(r'-?(before|after)-galactic-intervention$', '', loc)
    loc = re.sub(r'^(kanto|johto|hoenn|sinnoh|unova|kalos|alola|galar|paldea|hisui)-', '', loc)
    loc = loc.replace('sea-route-', 'Sea Route ').replace('route-', 'Route ')
    if '-' in loc or ' ' in loc:
        loc = ' '.join([w.capitalize() for w in re.split(r'[- ]', loc) if w])
    else:
        if loc: loc = loc.capitalize()
    return loc.strip()

def parse_chain(node):
    species_url = node['species']['url']
    species_id = int(species_url.strip('/').split('/')[-1])
    evolves_to = [parse_chain(n) for n in node.get('evolves_to', [])]
    details = []
    for d in node.get('evolution_details', []):
        details.append({
            'trigger': d['trigger']['name'] if d.get('trigger') else None,
            'min_level': d.get('min_level'),
            'item': d['item']['name'] if d.get('item') else None,
            'held_item': d['held_item']['name'] if d.get('held_item') else None,
            'min_happiness': d.get('min_happiness'),
            'time_of_day': d.get('time_of_day'),
            'known_move': d['known_move']['name'] if d.get('known_move') else None,
            'location': d['location']['name'] if d.get('location') else None,
        })
    return {
        'species_id': species_id,
        'is_baby': node.get('is_baby', False),
        'details': details,
        'evolves_to': evolves_to
    }

def get_gen_by_id(pid):
    if pid <= 151: return 1
    if pid <= 251: return 2
    if pid <= 386: return 3
    if pid <= 493: return 4
    if pid <= 649: return 5
    if pid <= 721: return 6
    if pid <= 809: return 7
    if pid <= 905: return 8
    return 9

def get_url(cell_value):
    if not isinstance(cell_value, str):
        return None
    match = re.search(r'(https?://[^\s"<>&\u0027\)]+)', cell_value)
    return match.group(1) if match else None

def get_form(name, is_direct=False):
    name = str(name).lower().replace('é', 'e').replace('è', 'e').strip()
    
    if is_direct:
        return re.sub(r'\s+', '-', name).replace("'", "")
        
    if 'alola' in name: return 'alola'
    if 'galar' in name: return 'galar'
    if 'hisui' in name: return 'hisui'
    if 'paldea' in name: return 'paldea'
    
    name_check = name.replace('*', '').strip()
    words = name_check.replace('(', ' ').replace(')', ' ').split()
    if 'mega' in words:
        if 'x' in words or 'x)' in name_check: return 'mega-x'
        if 'y' in words or 'y)' in name_check: return 'mega-y'
        return 'mega'
        
    if 'primal' in name: return 'primal'
    if 'gmax' in name or 'gigantamax' in name: return 'gmax'
    
    if '(' in name:
        form_part = name.split('(')[1].replace(')', '').strip()
        form_part = form_part.lower().replace('é', 'e').replace('è', 'e')
        
        base_forms = [
            'no plate', 'altered', 'land', 'incarnate', 'aria', 'ordinary',
            'shield', '50%', 'confined', 'baile', 'midday', 'solo', 'red core',
            'disguised', 'amped', 'ice face', 'full belly', 'single strike',
            'hero', 'chest', 'family of four', 'green plumage', 'curly', 'plant', 'west',
            'spring', 'normal', 'base', 'standard', 'red-striped', 'red meteor', 'red',
            'natural'
        ]
        if form_part in base_forms:
            return 'normal'
            
        return re.sub(r'\s+', '-', form_part).replace("'", "")
        
    return 'normal'

def to_key(name):
    return name.lower().replace('é', 'e').replace('è', 'e').replace(' ', '_').strip()

def main():
    os.makedirs('assets/db', exist_ok=True)
    conn = sqlite3.connect('assets/db/pokedex.sqlite')
    c = conn.cursor()

    print("Erstelle Tabellen...")
    c.executescript('''
        CREATE TABLE IF NOT EXISTS pokemon (
            id INTEGER PRIMARY KEY, name_de TEXT, name_en TEXT,
            has_gender_differences INTEGER, gender_rate INTEGER, capture_rate INTEGER,
            evolution_chain_id INTEGER, egg_groups TEXT, weight REAL, speed INTEGER
        );
        CREATE TABLE IF NOT EXISTS forms (
            id INTEGER PRIMARY KEY AUTOINCREMENT, pokemon_id INTEGER, name TEXT,
            form_type TEXT, min_gen INTEGER, image_id INTEGER, types TEXT, exclusive_regions TEXT
        );
        CREATE TABLE IF NOT EXISTS encounters (
            pokemon_id INTEGER, gen TEXT, version TEXT, location_data TEXT
        );
        CREATE TABLE IF NOT EXISTS evolutions (
            chain_id INTEGER PRIMARY KEY, chain_json TEXT
        );
        CREATE TABLE IF NOT EXISTS dex_orders (
            dex_name TEXT, pokemon_id INTEGER, order_index INTEGER
        );
        CREATE TABLE IF NOT EXISTS special_dexes (
            dex_name TEXT, pokemon_id INTEGER
        );
        CREATE TABLE IF NOT EXISTS ball_urls (
            ball_name TEXT PRIMARY KEY, image_url TEXT
        );
        CREATE TABLE IF NOT EXISTS matching_balls (
            unique_id TEXT PRIMARY KEY, normal_balls TEXT, shiny_balls TEXT
        );
    ''')

    for table in ['pokemon', 'forms', 'encounters', 'evolutions', 'dex_orders', 'special_dexes', 'ball_urls', 'matching_balls']:
        c.execute(f'DELETE FROM {table}')

    print("\n1. Verarbeite Matching Pokeballs (Excel & JSON)...")
    actual_path = EXCEL_PATH
    if not os.path.exists(actual_path):
        for file in os.listdir('.'):
            if 'Legal_Matching' in file and file.endswith('.xlsx'):
                actual_path = file
                break
        for file in os.listdir('bin'):
            if 'Legal_Matching' in file and file.endswith('.xlsx'):
                actual_path = f"bin/{file}"
                break
    
    url_to_key = {}
    key_to_url = {}
    balls_database = {}

    if os.path.exists(actual_path):
        try:
            wb = openpyxl.load_workbook(actual_path, data_only=False)
            
            intro_sheet = wb['Intro']
            for row in intro_sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                for col in range(len(row) - 1):
                    url = get_url(row[col])
                    if url:
                        name = str(row[col+1]).replace('"', '').strip()
                        if 'ball' in name.lower():
                            key = to_key(name)
                            url_to_key[url] = key
                            key_to_url[key] = url

            url_to_key['https://i.imgur.com/eru43o1.png'] = 'strange_ball'
            key_to_url['strange_ball'] = 'https://i.imgur.com/eru43o1.png'
            url_to_key['https://i.imgur.com/aeqHLEh.png'] = 'cherish_ball'
            key_to_url['cherish_ball'] = 'https://i.imgur.com/aeqHLEh.png'

            for sheet_name in wb.sheetnames:
                if sheet_name in ['Intro', 'Vivillon', 'Alcremie']: continue
                sheet = wb[sheet_name]
                rows = list(sheet.iter_rows(values_only=True))
                
                for r in range(0, len(rows), 7):
                    if r >= len(rows): break
                    row_0 = rows[r]
                    for col in range(len(row_0) - 1):
                        id_val = str(row_0[col] or '').strip()
                        id_match = re.sub(r'[^0-9]', '', id_val)
                        if not id_match: continue
                        
                        poke_id = int(id_match)
                        name_str = str(row_0[col+1] or '').replace('*', '').strip()
                        form = get_form(name_str)
                        unique_id = f"{poke_id}_{form}".replace("'", "")
                        
                        normal_balls, shiny_balls = [], []
                        for offset in [1, 2, 3]:
                            if r + offset < len(rows) and col < len(rows[r+offset]):
                                url = get_url(rows[r+offset][col])
                                if url and url in url_to_key and url_to_key[url] not in normal_balls:
                                    normal_balls.append(url_to_key[url])
                        
                        for offset in [4, 5, 6]:
                            if r + offset < len(rows) and col < len(rows[r+offset]):
                                url = get_url(rows[r+offset][col])
                                if url and url in url_to_key and url_to_key[url] not in shiny_balls:
                                    shiny_balls.append(url_to_key[url])
                        
                        norm_str = ",".join(normal_balls) if normal_balls else "any_ball"
                        shin_str = ",".join(shiny_balls) if shiny_balls else "any_ball"
                        balls_database[unique_id] = {'normal': norm_str, 'shiny': shin_str}

            if 'Vivillon' in wb.sheetnames:
                viv_rows = list(wb['Vivillon'].iter_rows(values_only=True))
                viv_forms = ['meadow', 'icy-snow', 'polar', 'tundra', 'continental', 'garden', 'elegant', 'modern', 'marine', 'archipelago', 'high-plains', 'sandstorm', 'river', 'monsoon', 'savanna', 'sun', 'ocean', 'jungle', 'fancy', 'poke-ball']
                
                for r in range(len(viv_rows)):
                    row_data = viv_rows[r]
                    for c_idx in range(len(row_data)):
                        cell_val = str(row_data[c_idx] or '').lower().strip()
                        if not cell_val or len(cell_val) < 3 or 'example' in cell_val or 'note' in cell_val: continue
                        
                        detected = next((v for v in viv_forms if v.replace('-', ' ') in cell_val.replace('-', ' ') or v.replace('-', '') in cell_val.replace('-', '')), None)
                        if detected:
                            unique_id = f"666_{detected}"
                            normal_balls, shiny_balls = [], []
                            for r_offset in range(1, 4):
                                if r + r_offset < len(viv_rows):
                                    for c_offset in [c_idx-1, c_idx, c_idx+1]:
                                        if 0 <= c_offset < len(viv_rows[r + r_offset]):
                                            url = get_url(viv_rows[r + r_offset][c_offset])
                                            if url and url in url_to_key:
                                                if r_offset == 1 and url_to_key[url] not in normal_balls: normal_balls.append(url_to_key[url])
                                                elif r_offset >= 2 and url_to_key[url] not in shiny_balls: shiny_balls.append(url_to_key[url])
                            
                            norm_str = ",".join(normal_balls) if normal_balls else "any_ball"
                            shin_str = ",".join(shiny_balls) if shiny_balls else (norm_str if normal_balls else "any_ball")
                            if normal_balls or unique_id not in balls_database:
                                balls_database[unique_id] = {'normal': norm_str, 'shiny': shin_str}

            if 'Alcremie' in wb.sheetnames:
                alc_rows = list(wb['Alcremie'].iter_rows(values_only=True))
                for r in range(len(alc_rows)):
                    row_data = alc_rows[r]
                    if not row_data or not row_data[0]: continue
                    
                    cell_val = str(row_data[0]).lower().strip()
                    if not cell_val or len(cell_val) < 4 or 'note' in cell_val: continue
                    
                    form = None
                    if 'ruby' in cell_val and 'swirl' in cell_val: form = 'ruby-swirl'
                    elif 'ruby' in cell_val: form = 'ruby-cream'
                    elif 'caramel' in cell_val: form = 'caramel-swirl'
                    elif 'rainbow' in cell_val: form = 'rainbow-swirl'
                    elif 'vanilla' in cell_val: form = 'vanilla-cream'
                    elif 'matcha' in cell_val: form = 'matcha-cream'
                    elif 'mint' in cell_val: form = 'mint-cream'
                    elif 'lemon' in cell_val: form = 'lemon-cream'
                    elif 'salted' in cell_val: form = 'salted-cream'
                    
                    if form:
                        unique_id = f"869_{form}"
                        normal_balls, shiny_balls = [], []
                        
                        for c_idx in range(1, len(row_data)):
                            url = get_url(row_data[c_idx])
                            if url and url in url_to_key and url_to_key[url] not in normal_balls: normal_balls.append(url_to_key[url])
                            
                        if r + 1 < len(alc_rows):
                            for c_idx in range(1, len(alc_rows[r+1])):
                                url = get_url(alc_rows[r+1][c_idx])
                                if url and url in url_to_key and url_to_key[url] not in shiny_balls: shiny_balls.append(url_to_key[url])
                        
                        norm_str = ",".join(normal_balls) if normal_balls else "any_ball"
                        shin_str = ",".join(shiny_balls) if shiny_balls else "any_ball"
                        
                        if normal_balls or unique_id not in balls_database:
                            balls_database[unique_id] = {'normal': norm_str, 'shiny': shin_str}

        except Exception as e:
            print(f"Fehler beim Laden/Parsen der Excel-Datei: {e}")

    if os.path.exists(CUSTOM_JSON_PATH):
        try:
            with open(CUSTOM_JSON_PATH, 'r', encoding='utf-8') as f:
                custom_data = json.load(f)
                for unique_id, data in custom_data.items():
                    norm_str = ",".join(data.get('normal', [])) if data.get('normal') else "any_ball"
                    shin_str = ",".join(data.get('shiny', [])) if data.get('shiny') else "any_ball"
                    balls_database[unique_id] = {'normal': norm_str, 'shiny': shin_str}
        except Exception as e:
            print(f"Fehler beim Laden der Custom JSON: {e}")

    print("\n2. Hole Pokemon Daten (Basis, Formen, Encounters, Speed/Weight)...")
    custom_enc = {}
    if os.path.exists('bin/custom_encounters.json'):
        try:
            with open('bin/custom_encounters.json', 'r', encoding='utf-8') as f:
                custom_enc = json.load(f)
        except Exception as e:
            print(f"Fehler beim Laden von custom_encounters.json: {e}")

    expected_app_uids = set()

    for i in range(1, 1026):
        try:
            species = fetch_json(f"https://pokeapi.co/api/v2/pokemon-species/{i}")
            poke = fetch_json(f"https://pokeapi.co/api/v2/pokemon/{i}")
            if not species or not poke: continue

            has_gender_diff = species.get('has_gender_differences', False)

            name_de = name_en = "Unknown"
            for n in species.get('names', []):
                if n['language']['name'] == 'de': name_de = n['name']
                if n['language']['name'] == 'en': name_en = n['name']
            
            weight = poke.get('weight', 0) / 10.0
            speed = next((s['base_stat'] for s in poke.get('stats', []) if s['stat']['name'] == 'speed'), 0)
            egg_groups = ",".join([eg['name'] for eg in species.get('egg_groups', [])])
            evo_chain_id = int(species['evolution_chain']['url'].strip('/').split('/')[-1]) if species.get('evolution_chain') else -1
            
            c.execute('''INSERT INTO pokemon
                (id, name_de, name_en, has_gender_differences, gender_rate, capture_rate, evolution_chain_id, egg_groups, weight, speed)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (i, name_de, name_en, 1 if has_gender_diff else 0,
                 species.get('gender_rate', -1), species.get('capture_rate', 255), evo_chain_id, egg_groups, weight, speed))

            varieties = species.get('varieties', [])
            if not varieties:
                expected_app_uids.add(f"{i}_normal")

            for variety in varieties:
                v_poke = fetch_json(variety['pokemon']['url'])
                if not v_poke: continue
                v_id = v_poke.get('id', i)
                
                forms_list = v_poke.get('forms', [])
                if not forms_list:
                    expected_app_uids.add(f"{i}_normal")

                for form_obj in forms_list:
                    f_data = fetch_json(form_obj['url'])
                    if not f_data: continue
                    
                    types = ",".join([t['type']['name'] for t in f_data.get('types', [])] if f_data.get('types') else [t['type']['name'] for t in v_poke.get('types', [])])
                    raw_name = f_data.get('name', '')
                    clean_form = raw_name.replace(species['name'], '').lstrip('-').strip() or 'normal'
                    expected_app_uids.add(f"{i}_{clean_form}")
                    
                    form_type = 'other'
                    min_gen = 9
                    if f_data.get('version_group'):
                        gen_str = VERSION_TO_GEN.get(f_data['version_group']['name'], 'gen_9').split('_')[-1]
                        min_gen = int(gen_str) if gen_str.isdigit() else 9
                    
                    if clean_form == 'normal': 
                        form_type = 'normal'
                        min_gen = get_gen_by_id(i)
                    elif any(x in clean_form for x in ['alola', 'galar', 'hisui', 'paldea']): form_type = 'regional'
                    elif 'mega' in clean_form or 'primal' in clean_form: form_type = 'mega'
                    elif 'gmax' in clean_form: form_type = 'gmax'
                    
                    exclusives = ",".join(FORM_WHITELIST.get(f"{i}_{clean_form}", []))
                    c.execute('''INSERT INTO forms (pokemon_id, name, form_type, min_gen, image_id, types, exclusive_regions)
                                 VALUES (?, ?, ?, ?, ?, ?, ?)''',
                              (i, clean_form, form_type, min_gen, v_id, types, exclusives))

            if has_gender_diff:
                expected_app_uids.add(f"{i}_m")
                expected_app_uids.add(f"{i}_f")

            encounters = fetch_json(f"https://pokeapi.co/api/v2/pokemon/{i}/encounters") or []
            enc_data = {}
            
            if str(i) in custom_enc:
                for gen, vmap in custom_enc[str(i)].items(): enc_data[gen] = vmap
                
            for enc in encounters:
                raw_loc = enc['location_area']['name']
                clean_loc = clean_location(raw_loc)
                
                for v_detail in enc.get('version_details', []):
                    version = v_detail['version']['name']
                    if '-japan' in version: continue
                    gen = VERSION_TO_GEN.get(version, 'gen_unknown')
                    
                    chance_agg = {}
                    for detail in v_detail.get('encounter_details', []):
                        method = detail['method']['name']
                        min_l, max_l = detail['min_level'], detail['max_level']
                        lvl_str = str(min_l) if min_l == max_l else f"{min_l}-{max_l}"
                        key = f"{method}|||{lvl_str}"
                        chance_agg[key] = chance_agg.get(key, 0) + detail['chance']
                        
                    for key, total_chance in chance_agg.items():
                        method, lvl_str = key.split('|||')
                        final_loc = f"{clean_loc}|||{method}|||{lvl_str}|||{total_chance}"
                        
                        if gen not in enc_data: enc_data[gen] = {}
                        if version not in enc_data[gen]: enc_data[gen][version] = []
                        if final_loc not in enc_data[gen][version]: enc_data[gen][version].append(final_loc)

            for gen, vmap in enc_data.items():
                for version, locs in vmap.items():
                    locs_str = "|||||".join(locs)
                    c.execute('INSERT INTO encounters (pokemon_id, gen, version, location_data) VALUES (?, ?, ?, ?)', (i, gen, version, locs_str))

            if species.get('is_legendary'): c.execute('INSERT INTO special_dexes (dex_name, pokemon_id) VALUES (?, ?)', ('legendary-dex', i))
            if species.get('is_mythical'): c.execute('INSERT INTO special_dexes (dex_name, pokemon_id) VALUES (?, ?)', ('mythical-dex', i))
            for eg in species.get('egg_groups', []): c.execute('INSERT INTO special_dexes (dex_name, pokemon_id) VALUES (?, ?)', (f"egg-{eg['name']}", i))
            
            if i % 50 == 0:
                print(f"  ... {i}/1025 bearbeitet")
                conn.commit()
                
        except Exception as e:
            print(f"Fehler bei Pokemon ID {i}: {e}")
            
    conn.commit()

    print("\n3. Führe API Form/Geschlechter-Fallbacks durch...")
    for uid in expected_app_uids:
        if uid not in balls_database or balls_database[uid]['normal'] == "any_ball":
            parts = uid.split('_', 1)
            if not parts[0].isdigit(): continue
            poke_id = int(parts[0])
            
            fallback_data = None
            base_key = f"{poke_id}_normal"
            
            if base_key in balls_database and balls_database[base_key]['normal'] != "any_ball":
                fallback_data = balls_database[base_key]
            else:
                for db_uid, db_data in balls_database.items():
                    if db_uid.startswith(f"{poke_id}_") and db_data['normal'] != "any_ball":
                        fallback_data = db_data
                        break
            
            if fallback_data:
                balls_database[uid] = {'normal': fallback_data['normal'], 'shiny': fallback_data['shiny']}
            else:
                balls_database[uid] = {'normal': 'any_ball', 'shiny': 'any_ball'}

    for poke_id in [493, 773]:
        for type_name, ball_key in TYPE_BALL_MAPPING.items():
            balls_database[f"{poke_id}_{type_name}"] = {'normal': ball_key, 'shiny': ball_key}
            if type_name == 'normal': balls_database[f"{poke_id}_normal"] = {'normal': ball_key, 'shiny': ball_key}

    print("Schreibe Matching Balls in die Datenbank...")
    for key, url in key_to_url.items():
        c.execute("INSERT OR IGNORE INTO ball_urls (ball_name, image_url) VALUES (?, ?)", (key, url))

    for uid, data in balls_database.items():
        c.execute("INSERT OR REPLACE INTO matching_balls (unique_id, normal_balls, shiny_balls) VALUES (?, ?, ?)", (uid, data['normal'], data['shiny']))
    conn.commit()

    print("\n4. Hole Evolutionsketten (1-550)...")
    for i in range(1, 551):
        try:
            chain_data = fetch_json(f"https://pokeapi.co/api/v2/evolution-chain/{i}")
            if chain_data and chain_data.get('chain'):
                parsed_chain = parse_chain(chain_data['chain'])
                c.execute('INSERT INTO evolutions (chain_id, chain_json) VALUES (?, ?)', (i, json.dumps(parsed_chain)))
            if i % 50 == 0: print(f"  ... {i}/550 geprueft")
        except Exception as e:
            print(f"Fehler bei Evolutionskette ID {i}: {e}")
            
    conn.commit()

    print("\n5. Lade Pokedex-Reihenfolgen...")
    try:
        for dex_name, map_key in REGIONAL_ENDPOINTS.items():
            if dex_name in HARDCODED_ORDERS:
                for idx, pid in enumerate(HARDCODED_ORDERS[dex_name]):
                    c.execute('INSERT INTO dex_orders (dex_name, pokemon_id, order_index) VALUES (?, ?, ?)', (map_key, pid, idx))
            else:
                dex_data = fetch_json(f"https://pokeapi.co/api/v2/pokedex/{dex_name}")
                if dex_data and dex_data.get('pokemon_entries'):
                    for idx, entry in enumerate(dex_data['pokemon_entries']):
                        url_parts = entry['pokemon_species']['url'].strip('/').split('/')
                        pid = int(url_parts[-1])
                        c.execute('INSERT INTO dex_orders (dex_name, pokemon_id, order_index) VALUES (?, ?, ?)', (map_key, pid, idx))
                        
        national_max = {'kanto': 151, 'johto': 251, 'hoenn': 386, 'sinnoh': 493, 'unova': 649, 'kalos': 721, 'alola': 809, 'galar': 905, 'paldea': 1025}
        for region, mx in national_max.items():
            for i in range(1, mx + 1):
                c.execute('INSERT INTO dex_orders (dex_name, pokemon_id, order_index) VALUES (?, ?, ?)', (f"{region}_national", i, i-1))
    except Exception as e:
        print(f"Fehler beim Laden der Dex-Reihenfolgen: {e}")

    conn.commit()
    conn.close()
    print("\nFertig! Datenbank 'assets/db/pokedex.sqlite' wurde inkl. Matching Balls erfolgreich erstellt!")

if __name__ == "__main__":
    main()